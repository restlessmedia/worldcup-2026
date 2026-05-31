"""Export frozen draw data from archive/draw-results.xlsx into data/*.json."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
ARCHIVE = ROOT / "archive" / "draw-results.xlsx"
DATA = ROOT / "data"


def read_sheet_rows(ws, max_rows: int | None = None) -> list[tuple]:
    limit = max_rows or ws.max_row
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=limit, values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            rows.append(tuple(cell for cell in row))
    return rows


def export_lists(ws) -> list[dict]:
    teams = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        canonical, common, display = (row + (None, None, None))[:3]
        if not canonical:
            continue
        teams.append(
            {
                "id": str(canonical).strip(),
                "common_name": str(common).strip() if common else None,
                "display_name": str(display).strip() if display else str(canonical).strip(),
            }
        )
    return teams


def export_responses(ws) -> list[dict]:
    houses = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        name, house_number, min_expected, requested, allocated, reduced, label = (
            row + (None,) * 7
        )[:7]
        if not name or not house_number:
            continue
        houses.append(
            {
                "respondent": str(name).strip(),
                "house_id": str(house_number).strip(),
                "min_expected_tickets": min_expected,
                "requested_tickets": requested,
                "allocated_tickets": allocated,
                "allocation_reduced": str(reduced).strip() if reduced else None,
                "label": str(label).strip() if label else None,
            }
        )
    return houses


def export_tracking(ws) -> list[dict]:
    draw = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        house_id, t1, t2, t3, t4, to_pay, paid, picker = (row + (None,) * 8)[:8]
        if house_id is None:
            continue
        teams = [t for t in (t1, t2, t3, t4) if t]
        if not teams:
            continue
        draw.append(
            {
                "house_id": str(house_id).strip(),
                "teams": [str(t).strip() for t in teams],
                "ticket_picker": str(picker).strip() if picker else None,
                "payment": {"to_pay": to_pay, "paid": paid},
            }
        )
    return draw


def export_config(data_ws, rules_ws) -> dict:
    prizes = []
    for row in data_ws.iter_rows(min_row=11, max_row=16, values_only=True):
        name, _, pct, amount = (row + (None, None, None, None))[:4]
        if not name:
            continue
        prizes.append(
            {
                "name": str(name).strip(),
                "percent": pct,
                "amount_gbp": amount,
            }
        )

    tie_rule = None
    for row in rules_ws.iter_rows(min_row=1, max_row=rules_ws.max_row, values_only=True):
        if row[0] and row[1] and str(row[0]).strip() != "Table 1":
            tie_rule = {
                "prize": str(row[0]).strip(),
                "order": str(row[1]).strip(),
            }
            break

    totals = {}
    for row in data_ws.iter_rows(min_row=3, max_row=9, values_only=True):
        label, value = (row + (None, None))[:2]
        if label and value is not None:
            key = (
                str(label)
                .strip()
                .lower()
                .replace(" ", "_")
                .replace("%", "pct")
                .rstrip("_")
            )
            totals[key] = value

    return {
        "tournament": "FIFA World Cup 2026",
        "currency": "GBP",
        "totals": totals,
        "prizes": prizes,
        "side_prize_tie_break": tie_rule,
        "scoring_model": "last_team_standing",
        "notes": [
            "Main prizes go to the house whose last remaining team reaches each stage.",
            "Side prizes: most goals conceded, fair play award.",
        ],
    }


def main() -> None:
    if not ARCHIVE.exists():
        raise FileNotFoundError(f"Missing archive workbook: {ARCHIVE}")

    wb = openpyxl.load_workbook(ARCHIVE, data_only=True)
    DATA.mkdir(parents=True, exist_ok=True)

    team_list = export_lists(wb["Lists"])
    draw = export_tracking(wb["Tracking"])
    responses = export_responses(wb["Responses"])
    config = export_config(wb["Data"], wb["Rules"])

    provenance = {
        "source_file": "archive/draw-results.xlsx",
        "exported_on": date.today().isoformat(),
        "status": "frozen",
        "description": "Line-in-the-sand snapshot of the sweepstake draw. Do not edit the workbook; update data/results.json during the tournament.",
    }

    outputs = {
        "provenance.json": provenance,
        "teams-list.json": team_list,
        "draw.json": draw,
        "responses.json": responses,
        "config.json": config,
    }

    for filename, payload in outputs.items():
        path = DATA / filename
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(f"Wrote {path}")

    print(f"Exported {len(draw)} houses, {len(team_list)} teams.")


if __name__ == "__main__":
    main()
